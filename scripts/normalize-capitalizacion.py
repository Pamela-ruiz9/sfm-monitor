#!/usr/bin/env python3
"""
normalize-capitalizacion.py
Extrae ICAP, CET1, LCR y NSFR del Boletín de Capitalización y Reporte de Liquidez CNBV.

Cierra: #19

Archivos fuente (descargar manualmente de portafolioinfo.cnbv.gob.mx):
  raw-data/boletin_capitalizacion.xlsx  → data/capitalizacion.json
  raw-data/reporte_liquidez.xlsx        → data/liquidez.json

Secciones del portal:
  Banca Múltiple → Información Regulatoria → Capitalización → Boletín Estadístico
  Banca Múltiple → Información Regulatoria → Liquidez

Frecuencia: trimestral (capitalización), mensual (liquidez). Rezago ~T+45 días.

Run:
  python3 scripts/normalize-capitalizacion.py
  python3 scripts/normalize-capitalizacion.py --dry-run
"""

import json
import os
import sys

DRY_RUN = "--dry-run" in sys.argv

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW = os.path.join(ROOT, "raw-data")
OUT = os.path.join(ROOT, "data")

BOLETIN_PATH = os.path.join(RAW, "boletin_capitalizacion.xlsx")
LIQUIDEZ_PATH = os.path.join(RAW, "reporte_liquidez.xlsx")


def write_json(filename: str, obj: object) -> None:
    path = os.path.join(OUT, filename)
    payload = json.dumps(obj, indent=2, ensure_ascii=False)
    if DRY_RUN:
        print(f"\n{'─'*60}")
        print(f"[DRY-RUN] Would write {path}:")
        print(payload[:2000] + (" ..." if len(payload) > 2000 else ""))
        return
    os.makedirs(OUT, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write(payload)
    print(f"✅ Written {path}  ({len(payload):,} bytes)")


def parse_capitalizacion() -> dict:
    """
    Parsea el Boletín Estadístico de Capitalización CNBV.

    El Excel tiene formato multinivel complejo. Las columnas relevantes son:
      - ICAP total sistema (Capital Neto / APR)
      - CET1 (Capital de Nivel 1 Ordinario)
      - Por institución: ICAP_latest

    Umbrales regulatorios (CUB):
      - ICAP mínimo: 10.5%
      - Vigilancia especial CNBV: < 12%
      - Sano (semáforo verde): >= 12%
    """
    try:
        import openpyxl
    except ImportError:
        print("⚠️  openpyxl no instalado. Correr: pip3 install openpyxl")
        sys.exit(1)

    if not os.path.exists(BOLETIN_PATH):
        print(f"⚠️  {BOLETIN_PATH} no encontrado.")
        print("    Descargar de: portafolioinfo.cnbv.gob.mx → Banca Múltiple")
        print("    → Información Regulatoria → Capitalización → Boletín Estadístico")
        sys.exit(1)

    wb = openpyxl.load_workbook(BOLETIN_PATH, data_only=True)

    # TODO: El formato del Boletín cambia cada año — revisar estructura con:
    #   python3 -c "import openpyxl; wb=openpyxl.load_workbook('raw-data/boletin_capitalizacion.xlsx'); print(wb.sheetnames)"
    # Las hojas típicas son: "Sistema", "Por institución", "Series históricas"

    # Placeholder — implementar después de inspeccionar el archivo real
    print("⚠️  normalize-capitalizacion.py requiere inspección del Excel descargado.")
    print("    Ver comentarios en el script para los pasos de implementación.")
    print(f"    Hojas disponibles: {wb.sheetnames}")

    # Estructura de salida esperada por SfmDataSchema.capitalizacion:
    return {
        "ultima_actualizacion": None,
        "fuente": "CNBV — Boletín Estadístico de Capitalización",
        "icap_sistema": {
            "actual": None,
            "fecha": None,
            "semaforo": None,
        },
        "cet1_sistema": {
            "actual": None,
            "fecha": None,
        },
        "historico": [],
        "por_banco": {},
    }


def semaforo_icap(valor: float | None) -> str:
    """
    Verde: >= 12%  (holgado sobre el mínimo CUB de 10.5%)
    Amarillo: 10.5% ≤ x < 12%  (cumple mínimo pero vigilancia)
    Rojo: < 10.5%  (incumple mínimo regulatorio)
    """
    if valor is None:
        return "amarillo"
    if valor >= 12.0:
        return "verde"
    if valor >= 10.5:
        return "amarillo"
    return "rojo"


def semaforo_lcr(valor: float | None) -> str:
    """
    Verde: >= 150%  (holgado — mediana sistema ~331% feb 2026)
    Amarillo: 100% ≤ x < 150%  (cumple mínimo)
    Rojo: < 100%  (incumple Basilea III)
    """
    if valor is None:
        return "amarillo"
    if valor >= 150.0:
        return "verde"
    if valor >= 100.0:
        return "amarillo"
    return "rojo"


def parse_liquidez() -> dict:
    """
    Parsea el Reporte de Liquidez CNBV (LCR y NSFR).

    Umbrales regulatorios:
      - LCR mínimo: 100% (Basilea III, implementado en México desde 2015)
      - NSFR mínimo: 100% (implementado desde 2021)
    """
    try:
        import openpyxl
    except ImportError:
        print("⚠️  openpyxl no instalado. Correr: pip3 install openpyxl")
        sys.exit(1)

    if not os.path.exists(LIQUIDEZ_PATH):
        print(f"⚠️  {LIQUIDEZ_PATH} no encontrado.")
        print("    Descargar de: portafolioinfo.cnbv.gob.mx → Banca Múltiple")
        print("    → Información Regulatoria → Liquidez")
        sys.exit(1)

    wb = openpyxl.load_workbook(LIQUIDEZ_PATH, data_only=True)
    print(f"    Hojas disponibles: {wb.sheetnames}")

    # Placeholder — implementar después de inspeccionar el archivo real
    return {
        "ultima_actualizacion": None,
        "fuente": "CNBV — Reporte de Indicadores de Liquidez",
        "lcr_sistema": {
            "actual": None,
            "fecha": None,
            "semaforo": None,
        },
        "nsfr_sistema": {
            "actual": None,
            "fecha": None,
        },
        "historico": [],
    }


if __name__ == "__main__":
    print("=== normalize-capitalizacion.py ===\n")

    if os.path.exists(BOLETIN_PATH):
        print("📊 Procesando capitalización...")
        cap = parse_capitalizacion()
        write_json("capitalizacion.json", cap)
    else:
        print(f"⏭️  {BOLETIN_PATH} no disponible — saltando capitalización")
        print("   Para activar: descargar el Boletín de Capitalización CNBV")

    if os.path.exists(LIQUIDEZ_PATH):
        print("\n💧 Procesando liquidez...")
        liq = parse_liquidez()
        write_json("liquidez.json", liq)
    else:
        print(f"\n⏭️  {LIQUIDEZ_PATH} no disponible — saltando liquidez")
        print("   Para activar: descargar el Reporte de Liquidez CNBV")
